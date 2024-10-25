import pandas as pd
from openpyxl import load_workbook
import os
import configparser

# Load the configuration
config = configparser.ConfigParser()
config.read('weekly-hours.ini')

folder_path = config.get('Paths', 'folder_path')
file_path = config.get('Paths', 'file_path')

employee_report_sheet = config.get('Sheets', 'employee_report_sheet')
weekly_forecast_sheet = config.get('Sheets', 'weekly_forecast_sheet')

employee_report_file_col1 = config.getint('Columns', 'employee_report_file_col1')
employee_report_file_col2 = config.getint('Columns', 'employee_report_file_col2')
weekly_file_col1 = config.getint('Columns', 'weekly_file_col1')
weekly_file_col2 = config.getint('Columns', 'weekly_file_col2')

start_row = config.getint('Processing', 'start_row')
end_row = config.getint('Processing', 'end_row')
col_offset = config.getint('Processing', 'col_offset')
row_offset = config.getint('Processing', 'row_offset')

employee_report_cols_start = config.getint('write', 'employee_report_cols_start')
employee_report_cols_end = config.getint('write', 'employee_report_cols_end')

# Load the main data file
df2 = pd.read_excel(file_path, sheet_name=employee_report_sheet)
data_from_second_file_col1 = df2.iloc[:, employee_report_file_col1].dropna().astype(str).str.strip().str.lower()
data_from_second_file_col2 = df2.iloc[:, employee_report_file_col2].dropna().astype(str).str.strip().str.lower()
df2['combined'] = data_from_second_file_col1 + "_" + data_from_second_file_col2
df2_combined = df2.dropna(subset=['combined'])

# Set of combined keys from the main data file
common_data = set(df2_combined['combined'])

# List to keep track of files with mismatches
mismatch_files = []

# Process each file in the folder
for file_name in os.listdir(folder_path):
    if file_name.endswith('.xlsx'):
        file1_path = os.path.join(folder_path, file_name)
        print(f"\nProcessing {file1_path}...")

        # Load the current file
        df1 = pd.read_excel(file1_path, sheet_name=weekly_forecast_sheet)
        data_from_first_file_col1 = df1.iloc[start_row:, weekly_file_col1].dropna().astype(str).str.strip().str.lower()
        data_from_first_file_col2 = df1.iloc[start_row:, weekly_file_col2].dropna().astype(str).str.strip().str.lower()
        df1['combined'] = data_from_first_file_col1 + "_" + data_from_first_file_col2
        df1_combined = df1.dropna(subset=['combined'])

        # Find common data between the current file and the main data file
        file_common_data = set(df1_combined['combined']) & common_data

        print("Common data found:")
        print(file_common_data)

        if file_common_data:
            print("Appending data...")
            workbook = load_workbook(file1_path)
            sheet = workbook[weekly_forecast_sheet]

            # Unmerge cells within the specified range
            for merged_cell in list(sheet.merged_cells.ranges):
                if merged_cell.min_row <= end_row and merged_cell.max_row >= start_row and merged_cell.min_col <= col_offset + employee_report_cols_end and merged_cell.max_col >= col_offset:
                    sheet.unmerge_cells(str(merged_cell))

            # Flag to track mismatches
            all_data_matched = True

            # Iterate through df1_combined to find matching rows and append data
            for idx, row in df1_combined.iterrows():
                combined_value = row['combined']
                if combined_value in file_common_data:
                    match_row = df2_combined[df2_combined['combined'] == combined_value].iloc[0]
                    for col_idx, value in enumerate(match_row.iloc[employee_report_cols_start:employee_report_cols_end]):
                        sheet.cell(row=row_offset, column=col_offset + col_idx, value=value)
                    print(f"Writing data for {combined_value} at row {row_offset}")
                    row_offset += 1  # Increment row_offset only for matched rows
                else:
                    print(f"{combined_value} not found in {file_path} data.")
                    all_data_matched = False  # Set flag to False if mismatch found

            # Save the workbook only if all data matched
            if all_data_matched:
                workbook.save(file1_path)
                print(f"Appended data successfully saved to {file1_path}")
            else:
                mismatch_files.append(file1_path)
                print(f"Mismatch found in {file1_path}. Process continues with next file.")
        else:
            print("No common data found in this file.")
            mismatch_files.append(file1_path)

if mismatch_files:
    print("\nMismatch found in the following files:")
    for mismatch_file in mismatch_files:
        print(mismatch_file)
else:
    print("\nProcessing completed without mismatches.")
