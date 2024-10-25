import pandas as pd
from openpyxl import load_workbook
import os
import configparser
config = configparser.ConfigParser()
config_file_path = 'consolidation.ini'  
config.read(config_file_path)
folder_path = config['Paths']['folder_path']
source_sheet = config['Sheets']['source_sheet']
source1_sheet = config['Sheets']['source1_sheet']
consolidated_file_path = os.path.join(folder_path, config['Paths']['consolidated_file_name'])
target_sheet_name = config['Sheets']['target_sheet_name']
first_empty_row = int(config['Sheets']['first_empty_row'])
weekly_forecast_column = int(config['Sheets']['weekly_forecast_column'])
project_data_start_row = int(config['IlocIndexes']['project_data_start_row'])
project_data_end_row = int(config['IlocIndexes']['project_data_end_row'])
project_data_start_col = int(config['IlocIndexes']['project_data_start_col'])
project_data_end_col = int(config['IlocIndexes']['project_data_end_col'])
base_forecast_revenue_start_row = int(config['IlocIndexes']['base_forecast_revenue_start_row'])
base_forecast_revenue_end_row = int(config['IlocIndexes']['base_forecast_revenue_end_row'])
base_forecast_revenue_start_col = int(config['IlocIndexes']['base_forecast_revenue_start_col'])
base_forecast_revenue_end_col = int(config['IlocIndexes']['base_forecast_revenue_end_col'])
weekly_forecast_revenue_start_row = int(config['IlocIndexes']['weekly_forecast_revenue_start_row'])
weekly_forecast_revenue_end_row = int(config['IlocIndexes']['weekly_forecast_revenue_end_row'])
weekly_forecast_revenue_start_col = int(config['IlocIndexes']['weekly_forecast_revenue_start_col'])
weekly_forecast_revenue_end_col = int(config['IlocIndexes']['weekly_forecast_revenue_end_col'])

df_list_1 = []
df_list_2 = []

print("Starting to process files...")

for file_name in os.listdir(folder_path):
    if file_name.endswith(".xlsx") and file_name != config['Paths']['consolidated_file_name']:
        file_path = os.path.join(folder_path, file_name)
        print(f"Processing file: {file_path}")
        
        try:
            data = pd.read_excel(file_path, sheet_name=source_sheet, header=None)
            extracted_data_1 = data.iloc[project_data_start_row:project_data_end_row, project_data_start_col:project_data_end_col].reset_index(drop=True)
            extracted_data_2 = data.iloc[base_forecast_revenue_start_row:base_forecast_revenue_end_row, base_forecast_revenue_start_col:base_forecast_revenue_end_col].reset_index(drop=True)
            
            data = pd.read_excel(file_path, sheet_name=source1_sheet, header=None)
            extracted_data_3 = data.iloc[weekly_forecast_revenue_start_row:weekly_forecast_revenue_end_row, weekly_forecast_revenue_start_col:weekly_forecast_revenue_end_col].reset_index(drop=True)
            
            df_list_1.append(pd.concat([extracted_data_1, extracted_data_2], axis=1))
            df_list_2.append(extracted_data_3)
        except Exception as e:
            print(f"Failed to process file {file_name}: {e}")

print("Combining data frames...")

combined_df_1 = pd.concat(df_list_1, ignore_index=True)
combined_df_2 = pd.concat(df_list_2, ignore_index=True)
print("Combined data frames created.")

try:
    workbook = load_workbook(consolidated_file_path)
    if target_sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(target_sheet_name)
        print(f"Created new sheet: {target_sheet_name}")
    else:
        sheet = workbook[target_sheet_name]
        while any(sheet.cell(row=first_empty_row, column=col).value is not None for col in range(1, sheet.max_column + 1)):
            first_empty_row += 1

    print(f"Writing data starting from row: {first_empty_row}")

    for r_idx, row in combined_df_1.iterrows():
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=first_empty_row + r_idx, column=c_idx, value=value)
    for r_idx, row in combined_df_2.iterrows():
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=first_empty_row + r_idx, column=weekly_forecast_column + c_idx - 1, value=value)

    workbook.save(consolidated_file_path)
    print(f"Consolidated data appended to {consolidated_file_path}")
except Exception as e:
    print(f"Failed to write to workbook: {e}")
 

 




